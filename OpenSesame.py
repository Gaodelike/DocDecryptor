import os
import logging
import shutil
from pathlib import Path
from typing import Union, Dict, Callable, Optional

# 动态导入 Office 文件处理库，避免在没有安装的情况下报错
try:
    from docx import Document
except ImportError:
    Document = None
    logging.warning("python-docx not installed. DOCX processing will be skipped.")

try:
    from PyPDF2 import PdfReader, PdfWriter
except ImportError:
    PdfReader = None
    PdfWriter = None
    logging.warning("PyPDF2 not installed. PDF processing will be skipped.")

try:
    from PIL import Image, UnidentifiedImageError
except ImportError:
    Image = None
    UnidentifiedImageError = None
    logging.warning("Pillow not installed. Image processing will be skipped.")

try:
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:
    load_workbook = None
    InvalidFileException = None
    logging.warning("openpyxl not installed. XLSX processing will be skipped.")

try:
    from pptx import Presentation
except ImportError:
    Presentation = None
    logging.warning("python-pptx not installed. PPTX processing will be skipped.")

import csv

# 配置日志（这里可以配置到文件，或者直接打印，根据GUI的日志显示需求调整）
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

BUFFER_SIZE = 4096  # 缓冲区大小，用于流式读取，优化大文件处理

# --- 加密标志常量 ---
ESAFENET_MARKER = b"Esafenet"
ESAFENET_SEARCH_CHUNK_SIZE = 1024  # 建议读取1024字节来查找标志


# --- 解密相关功能 ---
def dummy_decrypt_function(encrypted_data: bytes) -> bytes:
    """
    一个模拟的解密函数。
    实际应用中，这里会集成真正的解密算法和密钥。
    现在它将替换 Esafenet 标志为 DecryptedData_Marker。
    """
    logging.info("Simulating decryption process...")
    return encrypted_data.replace(ESAFENET_MARKER, b"DecryptedData_Marker")


def _copy_binary_file(src_path: Path, dest_path: Path, decrypt_func: Optional[Callable[[bytes], bytes]] = None):
    """
    以二进制流式方式复制文件，并可选地在复制过程中进行解密。
    dest_path 应该是完整的目标路径。
    """
    try:
        # 确保 dest_path 的父目录存在
        dest_path.parent.mkdir(parents=True, exist_ok=True)

        with open(src_path, "rb") as src_file:
            with open(dest_path, "wb") as dest_file:
                while True:
                    chunk = src_file.read(BUFFER_SIZE)
                    if not chunk:
                        break
                    if decrypt_func:
                        processed_chunk = decrypt_func(chunk)
                        dest_file.write(processed_chunk)
                    else:
                        dest_file.write(chunk)
        logging.info(f"Successfully copied (and optionally processed) file: {src_path} to {dest_path}")
    except IOError as e:
        logging.error(f"Error copying file {src_path} to {dest_path}: {e}")
        raise


# --- 加密判断函数 ---
def is_esafenet_encrypted(file_path: Path) -> bool:
    """
    判断文件开头一定范围的字节内是否包含 Esafenet ASCII 字符串作为加密标识。
    """
    try:
        if not file_path.exists() or not file_path.is_file():
            logging.debug(f"File not found or not a regular file for encryption check: {file_path}")
            return False

        with open(file_path, "rb") as f:
            # 读取 ESAFENET_SEARCH_CHUNK_SIZE 字节来查找标志
            header_chunk = f.read(ESAFENET_SEARCH_CHUNK_SIZE)

            if ESAFENET_MARKER in header_chunk:
                logging.info(
                    f"Detected Esafenet encryption marker within first {ESAFENET_SEARCH_CHUNK_SIZE} bytes of file: {file_path}")
                return True
            else:
                logging.debug(
                    f"Esafenet marker not found within first {ESAFENET_SEARCH_CHUNK_SIZE} bytes of file: {file_path}")
                return False
    except IOError as e:
        logging.error(f"Error accessing file {file_path} for encryption check: {e}")
        return False
    except Exception as e:
        logging.error(f"An unexpected error occurred during encryption marker detection for {file_path}: {e}")
        return False


# --- 各类文件处理函数，统一后缀为 .info ---
# 注意：所有生成 new_file_path 的地方都需要修改
def read_and_save_docx(file_path: Path, output_dir: Path) -> Path:
    if not Document: raise ImportError("python-docx is not installed or import failed.")
    try:
        new_file_path = output_dir / f"{file_path.stem}.info"  # <<< 修改点：统一后缀为 .info
        new_file_path.parent.mkdir(parents=True, exist_ok=True)
        doc = Document(file_path)
        doc.save(new_file_path)
        logging.info(f"Successfully processed DOCX: {file_path} to {new_file_path}")
        return new_file_path
    except Exception as e:
        logging.error(f"Error processing DOCX {file_path}: {e}")
        raise


def read_and_save_pdf(file_path: Path, output_dir: Path) -> Path:
    if not PdfReader: raise ImportError("PyPDF2 is not installed or import failed.")
    try:
        new_file_path = output_dir / f"{file_path.stem}.info"  # <<< 修改点：统一后缀为 .info
        new_file_path.parent.mkdir(parents=True, exist_ok=True)
        reader = PdfReader(file_path)
        writer = PdfWriter()
        for page in reader.pages:
            writer.add_page(page)
        with open(new_file_path, "wb") as output_file:
            writer.write(output_file)
        logging.info(f"Successfully processed PDF: {file_path} to {new_file_path}")
        return new_file_path
    except Exception as e:
        logging.error(f"Error processing PDF {file_path}: {e}")
        raise


def read_and_save_image(file_path: Path, output_dir: Path) -> Path:
    if not Image: raise ImportError("Pillow is not installed or import failed.")
    try:
        new_file_path = output_dir / f"{file_path.stem}.info"  # <<< 修改点：统一后缀为 .info
        new_file_path.parent.mkdir(parents=True, exist_ok=True)
        image = Image.open(file_path)
        image.save(new_file_path)
        logging.info(f"Successfully processed Image: {file_path} to {new_file_path}")
        return new_file_path
    except UnidentifiedImageError as e:
        logging.error(f"Error: Not a valid image file {file_path}: {e}")
        raise
    except Exception as e:
        logging.error(f"Error processing Image {file_path}: {e}")
        raise


def read_and_save_xlsx(file_path: Path, output_dir: Path) -> Path:
    if not load_workbook: raise ImportError("openpyxl is not installed or import failed.")
    try:
        new_file_path = output_dir / f"{file_path.stem}.info"  # <<< 修改点：统一后缀为 .info
        new_file_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = load_workbook(file_path)
        workbook.save(new_file_path)
        logging.info(f"Successfully processed XLSX: {file_path} to {new_file_path}")
        return new_file_path
    except InvalidFileException as e:
        logging.error(f"Error: Not a valid Excel file {file_path}: {e}")
        raise
    except Exception as e:
        logging.error(f"Error processing XLSX {file_path}: {e}")
        raise


def read_and_save_pptx(file_path: Path, output_dir: Path) -> Path:
    if not Presentation: raise ImportError("python-pptx is not installed or import failed.")
    try:
        new_file_path = output_dir / f"{file_path.stem}.info"  # <<< 修改点：统一后缀为 .info
        new_file_path.parent.mkdir(parents=True, exist_ok=True)
        presentation = Presentation(file_path)
        presentation.save(new_file_path)
        logging.info(f"Successfully processed PPTX: {file_path} to {new_file_path}")
        return new_file_path
    except Exception as e:
        logging.error(f"Error processing PPTX {file_path}: {e}")
        raise


def read_and_save_csv(file_path: Path, output_dir: Path) -> Path:
    try:
        new_file_path = output_dir / f"{file_path.stem}.info"  # <<< 修改点：统一后缀为 .info
        new_file_path.parent.mkdir(parents=True, exist_ok=True)
        with open(file_path, newline='', encoding='utf-8') as infile:
            reader = csv.reader(infile)
            with open(new_file_path, mode='w', newline='', encoding='utf-8') as outfile:
                writer = csv.writer(outfile)
                for row in reader:
                    writer.writerow(row)
        logging.info(f"Successfully processed CSV: {file_path} to {new_file_path}")
        return new_file_path
    except Exception as e:
        logging.error(f"Error processing CSV {file_path}: {e}")
        raise


# --- 最终修改 process_file_unified 函数 ---
def process_file_unified(file_path_str: str, output_dir_str: str) -> Optional[Path]:
    """
    统一的文件处理入口。
    返回处理后的新文件路径，如果处理失败则返回 None。
    """
    file_path = Path(file_path_str)
    output_dir = Path(output_dir_str)
    file_extension = file_path.suffix.lower()

    if not file_path.exists():
        logging.error(f"File not found: {file_path}")
        return None

    # 确保输出目录存在
    output_dir.mkdir(parents=True, exist_ok=True)

    try:
        if file_extension == '.docx':
            return read_and_save_docx(file_path, output_dir)
        elif file_extension == '.pdf':
            return read_and_save_pdf(file_path, output_dir)
        elif file_extension in ['.jpg', '.jpeg', '.png', '.bmp', '.gif']:
            return read_and_save_image(file_path, output_dir)
        elif file_extension == '.xlsx':
            return read_and_save_xlsx(file_path, output_dir)
        elif file_extension == '.pptx':
            return read_and_save_pptx(file_path, output_dir)
        elif file_extension == '.csv':
            return read_and_save_csv(file_path, output_dir)
        else:
            # 对于未知文件类型，只检查 Esafenet 标志
            if is_esafenet_encrypted(file_path):
                logging.info(f"File {file_path} identified as Esafenet encrypted. Attempting decryption.")
                # 解密后的文件也统一后缀为 .info
                new_file_path = output_dir / f"{file_path.stem}_decrypted.info"  # <<< 修改点：统一后缀为 .info
                _copy_binary_file(file_path, new_file_path, decrypt_func=dummy_decrypt_function)
                return new_file_path
            else:
                # 如果没有找到 Esafenet 标志，则将其视为未知文件类型，并执行通用复制
                final_output_path = output_dir / f"{file_path.stem}_unhandled.info"  # <<< 修改点：统一后缀为 .info
                shutil.copy2(file_path, final_output_path)
                logging.warning(
                    f"File {file_path} is an unhandled type and no 'Esafenet' encryption marker found. Copied to {final_output_path}.")
                return final_output_path

    except ImportError as ie:
        logging.error(f"Missing required library for file type {file_extension}: {ie}")
        return None
    except Exception as e:
        logging.critical(f"A critical error occurred while processing {file_path}: {e}", exc_info=True)
        return None
